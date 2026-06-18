from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import io

SOLO_TAB_COMPANIES = {'YS Affiliates'}  # companies that get their own tab


def read_pl_sheet(ws):
    headers = None
    rows = []
    indents = {}  # label -> indent level from source
    seen_labels = {}  # label -> indent of first occurrence
    for row in ws.iter_rows():
        values = [c.value for c in row]
        if headers is None and isinstance(values[1], str) and 'Tickets' in str(values[1]):
            headers = values
            continue

        label_cell = row[0]
        label = label_cell.value
        if label is not None and isinstance(label, str):
            ind = 0
            if label_cell.alignment and label_cell.alignment.indent:
                ind = int(label_cell.alignment.indent)
            # Disambiguate a child account sharing its section header's name:
            # the first (shallower) occurrence keeps the name; a later deeper
            # occurrence is renamed to "<name> (detail)".
            if label in seen_labels and ind > seen_labels[label]:
                new_label = f"{label} (detail)"
                values = [new_label] + values[1:]
                label = new_label
            else:
                seen_labels[label] = ind
            if label not in indents:
                indents[label] = ind
        rows.append(values)
    return headers, rows, indents


def extract_ordered_rows(rows):
    result = []
    seen = set()
    for row in rows:
        label = row[0]
        if label is None:
            continue
        vals = row[1:]
        if label not in seen:
            result.append((label, vals))
            seen.add(label)
    return result


def merge_label_orders(primary_labels, secondary_labels):
    known = set(primary_labels)
    result = list(primary_labels)
    for i, label in enumerate(secondary_labels):
        if label in known:
            continue
        insert_after = None
        for j in range(i - 1, -1, -1):
            if secondary_labels[j] in known:
                insert_after = secondary_labels[j]
                break
        if insert_after and insert_after in result:
            pos = result.index(insert_after) + 1
            result.insert(pos, label)
        else:
            result.append(label)
        known.add(label)
    return result


def write_pl_sheet(ws, companies, all_labels, mar_map, apr_map,
                   mar_label, apr_label, year,
                   SKIP_ROWS, TOTAL_ROWS, SECTION_DIVIDERS, SECTION_LABEL_ONLY,
                   indents=None):
    """Write a P&L sheet for the given list of (orig_idx, name) companies."""
    if indents is None:
        indents = {}

    NAVY = "1F2D5A"
    TEAL = "006D6F"
    LTGRAY = "F2F3F5"
    WHITE = "FFFFFF"
    BLUE_HDR = "E3EAF5"
    TOTAL_BG = "C8D6EF"

    n_companies = len(companies)

    def company_start_col(ci):
        return 2 + ci * 5

    total_cols = 1 + n_companies * 5 - 1

    def row_style(label):
        if label == 'Gross Profit': return 'gross'
        if label == 'Net Operating Income': return 'noi'
        if label == 'Net Income': return 'ni'
        if label == 'Net Other Income': return 'netother'
        if label in TOTAL_ROWS: return 'total'
        if label in SECTION_DIVIDERS or label in SECTION_LABEL_ONLY: return 'section'
        return 'data'

    def has_any_data(label):
        mv = mar_map.get(label, [])
        av = apr_map.get(label, [])
        return (any(isinstance(v, (int, float)) and v != 0 for v in mv) or
                any(isinstance(v, (int, float)) and v != 0 for v in av))

    HEADER_ACCENT = "DCE3F0"   # slight blue accent for company names
    SUBHDR_ACCENT = "EEF1F6"   # lighter accent for the Mar/Apr/$/% row
    SPACER_GRAY = "CCCCCC"

    # Row 1: Company headers — slight accent fill, dark text
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=WHITE)
    for ci, (orig_idx, company) in enumerate(companies):
        cs = company_start_col(ci)
        ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=cs + 3)
        c = ws.cell(row=1, column=cs, value=company)
        c.font = Font(bold=True, size=8, color=NAVY, name="Arial")
        c.fill = PatternFill("solid", fgColor=HEADER_ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if ci < n_companies - 1:
            ws.cell(row=1, column=cs + 4).fill = PatternFill("solid", fgColor=SPACER_GRAY)
    ws.row_dimensions[1].height = 30

    # Column A row 1: comparison title (later month over earlier month)
    yy = str(year)[-2:]
    a1 = ws.cell(row=1, column=1,
                 value=f"{apr_label} '{yy} over {mar_label} '{yy} Comparison")
    a1.font = Font(bold=True, size=10, color=NAVY, name="Arial")
    a1.fill = PatternFill("solid", fgColor=HEADER_ACCENT)
    a1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Row 2: Sub-headers — light accent fill, dark text
    a2 = ws.cell(row=2, column=1, value="Account")
    a2.font = Font(bold=True, size=9, color=NAVY, name="Arial")
    a2.fill = PatternFill("solid", fgColor=HEADER_ACCENT)
    a2.alignment = Alignment(horizontal="left", vertical="center")
    for ci in range(n_companies):
        cs = company_start_col(ci)
        for j, hdr in enumerate([mar_label[:3], apr_label[:3], "$ Chg", "% Chg"]):
            c = ws.cell(row=2, column=cs + j, value=hdr)
            c.font = Font(bold=True, size=8, color=NAVY, name="Arial")
            c.fill = PatternFill("solid", fgColor=SUBHDR_ACCENT)
            c.alignment = Alignment(horizontal="center")
        if ci < n_companies - 1:
            ws.cell(row=2, column=cs + 4).fill = PatternFill("solid", fgColor=SPACER_GRAY)
    ws.row_dimensions[2].height = 15

    spacer_cols = {company_start_col(ci) + 4 for ci in range(n_companies - 1)}
    pct_cols = {company_start_col(ci) + 3 for ci in range(n_companies)}

    # Build set of labels that have at least one non-zero value for these companies.
    # Exclude 'Total' from this check — Total aggregates everything including solo-tab
    # companies, so it would cause rows to appear even when no entity on this tab has data.
    non_total_companies = [(i, name) for i, name in companies if name != 'Total']

    def company_has_data(label):
        mv = mar_map.get(label, [])
        av = apr_map.get(label, [])
        check_list = non_total_companies if non_total_companies else companies
        for orig_idx, _ in check_list:
            m = mv[orig_idx] if orig_idx < len(mv) else None
            a = av[orig_idx] if orig_idx < len(av) else None
            if (isinstance(m, (int, float)) and m != 0) or (isinstance(a, (int, float)) and a != 0):
                return True
        return False

    def indent_of(label):
        return indents.get(label, 0)

    # Identify parent rows that head a group (a row immediately followed by
    # deeper-indented rows). Works for both data-bearing parents (e.g. Commissions)
    # and label-only parents (e.g. Professional Fees).
    parents_with_children = set()
    for i, lbl in enumerate(all_labels):
        if str(lbl).startswith('Total for'):
            continue
        nxt = all_labels[i + 1] if i + 1 < len(all_labels) else None
        if nxt and indent_of(nxt) > indent_of(lbl):
            parents_with_children.add(lbl)

    # Map each parent to its contiguous child labels (deeper-indented, non-subtotal)
    children_of = {}
    for i, lbl in enumerate(all_labels):
        if lbl in parents_with_children:
            kids = []
            base = indent_of(lbl)
            for j in range(i + 1, len(all_labels)):
                if indent_of(all_labels[j]) > base and not str(all_labels[j]).startswith('Total for'):
                    kids.append(all_labels[j])
                else:
                    break
            children_of[lbl] = kids

    # Only always show the key P&L summary lines and structural section headers;
    # everything else filters by whether the company has data.
    ALWAYS_SHOW = {'Gross Profit', 'Net Operating Income', 'Net Income', 'Net Other Income'}

    def should_show(label):
        if label in ALWAYS_SHOW:
            return True
        # Section dividers (Income, Cost of Goods Sold, Expenses, Other Income, Other Expenses)
        if label in SECTION_DIVIDERS:
            return True
        # Parent rows that head a group: show if any child has data
        if label in parents_with_children:
            return any(company_has_data(c) for c in children_of.get(label, []))
        # Subtotals: show if their group parent's children have data
        if label in TOTAL_ROWS:
            parent = label.replace('Total for ', '', 1)
            if parent in parents_with_children:
                return any(company_has_data(c) for c in children_of.get(parent, []))
            return company_has_data(label)
        return company_has_data(label)

    data_start_row = 3
    used_rows = []
    gp_excel_row = None
    income_excel_row = None

    for label in all_labels:
        if label in SKIP_ROWS or label == 'Consolidated P&L':
            continue
        if not should_show(label):
            continue

        excel_row = data_start_row + len(used_rows)
        used_rows.append(label)
        style = row_style(label)
        ind = indent_of(label)

        if label == 'Gross Profit':
            gp_excel_row = excel_row
        if label == 'Total for Income':
            income_excel_row = excel_row

        # A real section divider is indent-0 (Income, Expenses, etc.) and is NOT
        # a group parent. Label-only parents like "Professional Fees" should be
        # styled as group parents, not section dividers.
        is_section_divider = (label in SECTION_DIVIDERS)
        is_group_parent = (label in parents_with_children) and not is_section_divider
        is_subaccount = (style == 'data' and ind >= 2)
        in_group = is_subaccount or is_group_parent

        # Unified palette: no greens, one blue for all subtotal/total rows,
        # one tint for group blocks, flat white for standalone rows.
        SUBTOTAL_BLUE = "C8D6EF"   # all "Total for X" and grand-total rows
        GROUP_TINT = "EEF1F6"      # parent + child group blocks
        SECTION_BLUE = "DCE3F0"    # section dividers (Income, Expenses, ...)
        GROUP_PARENT_FG = "1F2D5A"

        if style in ('gross', 'noi', 'ni', 'netother', 'total'):
            bg, fg = SUBTOTAL_BLUE, NAVY
        elif is_section_divider:
            bg, fg = SECTION_BLUE, "1F2D5A"
        elif is_group_parent:
            bg, fg = GROUP_TINT, GROUP_PARENT_FG
        elif is_subaccount:
            bg, fg = GROUP_TINT, "000000"
        elif style == 'section':
            bg, fg = SECTION_BLUE, "1F2D5A"
        else:
            bg, fg = WHITE, "000000"

        display_label = label[:-9] if label.endswith(' (detail)') else label
        lc = ws.cell(row=excel_row, column=1, value=display_label)
        # Section dividers and group parents are bold to anchor structure.
        lc.font = Font(name="Arial", size=9,
                       bold=(style != 'data' or is_group_parent), color=fg)
        lc.fill = PatternFill("solid", fgColor=bg)
        # Use the source indent level so the parent/child hierarchy is visible.
        if is_section_divider:
            label_indent = 0
        elif is_group_parent:
            label_indent = max(ind, 1)
        elif style == 'data':
            label_indent = max(ind, 1)
        elif style == 'total':
            label_indent = 1
        else:
            label_indent = 0
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=label_indent)

        # Subtotal rows ("Total for X") are computed by summing their group's
        # displayed child rows, so the subtotal always ties to what's shown.
        # Parent rows (group parents / section dividers) are left blank headers.
        is_subtotal = (label in TOTAL_ROWS)
        if is_subtotal:
            parent = label.replace('Total for ', '', 1)
            child_labels = children_of.get(parent, [])
        else:
            child_labels = []

        mar_vals = mar_map.get(label, [])
        apr_vals = apr_map.get(label, [])
        # Group parents and section dividers carry no own data (blank header).
        if is_group_parent or is_section_divider:
            write_data = False
        elif is_subtotal:
            write_data = True
        else:
            write_data = style != 'section' or has_any_data(label)

        def sum_children(value_map, idx):
            vals = [value_map.get(c, []) for c in child_labels]
            nums = [v[idx] for v in vals if idx < len(v) and isinstance(v[idx], (int, float))]
            return sum(nums) if nums else None

        # Pre-compute tab-scoped totals (sum of non-Total companies on this tab)
        non_total = [(i, n) for i, n in companies if n != 'Total']
        tab_mar_total = tab_apr_total = None
        if (write_data or is_subtotal) and non_total:
            if is_subtotal:
                m_each = [sum_children(mar_map, i) for i, _ in non_total]
                a_each = [sum_children(apr_map, i) for i, _ in non_total]
                m_each = [x for x in m_each if isinstance(x, (int, float))]
                a_each = [x for x in a_each if isinstance(x, (int, float))]
            else:
                m_each = [mar_vals[i] for i, _ in non_total
                          if i < len(mar_vals) and isinstance(mar_vals[i], (int, float))]
                a_each = [apr_vals[i] for i, _ in non_total
                          if i < len(apr_vals) and isinstance(apr_vals[i], (int, float))]
            if m_each: tab_mar_total = sum(m_each)
            if a_each: tab_apr_total = sum(a_each)

        for ci, (orig_idx, company) in enumerate(companies):
            cs = company_start_col(ci)
            if company == 'Total':
                mv = tab_mar_total
                av = tab_apr_total
            elif is_subtotal:
                mv = sum_children(mar_map, orig_idx)
                av = sum_children(apr_map, orig_idx)
            else:
                mv = mar_vals[orig_idx] if orig_idx < len(mar_vals) else None
                av = apr_vals[orig_idx] if orig_idx < len(apr_vals) else None
                mv = mv if isinstance(mv, (int, float)) else None
                av = av if isinstance(av, (int, float)) else None

            delta = pct = None
            if write_data:
                if mv is not None and av is not None:
                    delta = av - mv
                elif av is not None:
                    delta = av
                elif mv is not None:
                    delta = -mv
                pct = (delta / abs(mv)) if (mv is not None and mv != 0 and delta is not None) else None

            cell_bg = bg

            for j, (val, fmt) in enumerate([
                (mv if write_data else None, '#,##0;(#,##0);"-"'),
                (av if write_data else None, '#,##0;(#,##0);"-"'),
                (delta, '#,##0;(#,##0);"-"'),
                (pct, '0.00%;(0.00%);"-"'),
            ]):
                cell = ws.cell(row=excel_row, column=cs + j)
                cell.fill = PatternFill("solid", fgColor=cell_bg)
                cell.font = Font(name="Arial", size=9,
                                 bold=(style != 'data' or is_group_parent), color=fg)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    cell.value = val
                    cell.number_format = fmt

            if ci < n_companies - 1:
                ws.cell(row=excel_row, column=cs + 4).fill = PatternFill("solid", fgColor="CCCCCC")

        ws.row_dimensions[excel_row].height = 14

    # GPM row
    if gp_excel_row and income_excel_row:
        gpm_excel_row = gp_excel_row + 1
        ws.insert_rows(gpm_excel_row)
        bg_gpm, fg_gpm = "C8D6EF", "1F2D5A"
        lc = ws.cell(row=gpm_excel_row, column=1, value="Gross Profit Margin")
        lc.font = Font(name="Arial", size=9, bold=True, color=fg_gpm)
        lc.fill = PatternFill("solid", fgColor=bg_gpm)
        lc.alignment = Alignment(horizontal="left", vertical="center")

        for ci in range(n_companies):
            cs = company_start_col(ci)
            for j, src_col in enumerate([cs, cs + 1]):
                cell = ws.cell(row=gpm_excel_row, column=cs + j)
                cell.value = f"={get_column_letter(src_col)}{gp_excel_row}/{get_column_letter(src_col)}{income_excel_row}"
                cell.number_format = '0.00%;(0.00%);"-"'
                cell.font = Font(name="Arial", size=9, bold=True, color=fg_gpm)
                cell.fill = PatternFill("solid", fgColor=bg_gpm)
                cell.alignment = Alignment(horizontal="right", vertical="center")

            dc = ws.cell(row=gpm_excel_row, column=cs + 2)
            dc.fill = PatternFill("solid", fgColor=bg_gpm)
            dc.font = Font(name="Arial", size=9, bold=True, color=fg_gpm)

            pc = ws.cell(row=gpm_excel_row, column=cs + 3)
            pc.value = f"={get_column_letter(cs + 1)}{gpm_excel_row}-{get_column_letter(cs)}{gpm_excel_row}"
            pc.number_format = '0.00%;(0.00%);"-"'
            pc.font = Font(name="Arial", size=9, bold=True, color=fg_gpm)
            pc.fill = PatternFill("solid", fgColor=bg_gpm)
            pc.alignment = Alignment(horizontal="right", vertical="center")

            if ci < n_companies - 1:
                ws.cell(row=gpm_excel_row, column=cs + 4).fill = PatternFill("solid", fgColor="CCCCCC")

        ws.row_dimensions[gpm_excel_row].height = 14

    # Column widths
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            col = cell.column
            if col in spacer_cols:
                continue
            val = cell.value
            if val is None or (isinstance(val, str) and val.startswith('=')):
                continue
            if col in pct_cols and isinstance(val, float):
                rendered = f"({abs(val * 100):,.2f}%)" if val < 0 else f"{val * 100:,.2f}%"
            elif isinstance(val, (int, float)):
                rendered = f"({abs(val):,.0f})" if val < 0 else f"{val:,.0f}"
            else:
                rendered = str(val)
            col_widths[col] = max(col_widths.get(col, 0), len(rendered))

    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 3, 8), 40)
    for col in spacer_cols:
        ws.column_dimensions[get_column_letter(col)].width = 3

    ws.freeze_panes = "B3"


def build_combined_pl(mar_file, apr_file, mar_label="March", apr_label="April", year="2026"):
    mar_wb = load_workbook(mar_file, data_only=True)
    apr_wb = load_workbook(apr_file, data_only=True)
    mar_ws = mar_wb.active
    apr_ws = apr_wb.active

    mar_headers, mar_rows, mar_indents = read_pl_sheet(mar_ws)
    apr_headers, apr_rows, apr_indents = read_pl_sheet(apr_ws)

    # Merge indents (prefer March, fall back to April)
    indents = dict(apr_indents)
    indents.update(mar_indents)

    all_companies = mar_headers[1:]

    # Split into main companies and solo-tab companies
    main_companies = [(i, name) for i, name in enumerate(all_companies)
                      if name != 'Eliminations' and name not in SOLO_TAB_COMPANIES]
    solo_companies = [(i, name) for i, name in enumerate(all_companies)
                      if name in SOLO_TAB_COMPANIES]
    # Total always goes last on main sheet
    total_companies = [(i, name) for i, name in main_companies if name == 'Total']
    main_companies = [(i, name) for i, name in main_companies if name != 'Total'] + total_companies

    mar_ordered = extract_ordered_rows(mar_rows)
    apr_ordered = extract_ordered_rows(apr_rows)
    mar_map = {r[0]: r[1] for r in mar_ordered}
    apr_map = {r[0]: r[1] for r in apr_ordered}

    mar_labels = [r[0] for r in mar_ordered]
    apr_labels = [r[0] for r in apr_ordered]
    all_labels = merge_label_orders(mar_labels, apr_labels)

    SECTION_LABEL_ONLY = {'Miscellaneous Fees', 'Professional Fees', 'Salaries', 'K-1 Income'}
    SKIP_ROWS = {'Consolidated P&L', 'March 2026', 'April 2026',
                 f'{mar_label} {year}', f'{apr_label} {year}'}
    TOTAL_ROWS = {l for l in all_labels if l and str(l).startswith('Total for')}
    SECTION_DIVIDERS = {'Income', 'Cost of Goods Sold', 'Expenses', 'Other Income', 'Other Expenses'}

    shared_args = (all_labels, mar_map, apr_map, mar_label, apr_label, year,
                   SKIP_ROWS, TOTAL_ROWS, SECTION_DIVIDERS, SECTION_LABEL_ONLY, indents)

    wb = Workbook()

    # Main sheet
    ws_main = wb.active
    ws_main.title = "Combined P&L"
    write_pl_sheet(ws_main, main_companies, *shared_args)

    # Solo tabs
    for orig_idx, name in solo_companies:
        tab_title = name.replace(' Tickets', '').replace(' LLC', '').strip()[:31]
        ws_solo = wb.create_sheet(title=tab_title)
        write_pl_sheet(ws_solo, [(orig_idx, name)], *shared_args)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
