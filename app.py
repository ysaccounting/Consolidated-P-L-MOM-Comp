from flask import Flask, request, send_file, render_template, jsonify
import io
import os
import re
from openpyxl import load_workbook
from pl_builder import build_combined_pl

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def detect_month_year(file_bytes):
    """Read the month/year label from the header rows of the P&L sheet."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(max_row=5, values_only=True):
            val = row[0]
            if val and isinstance(val, str):
                for month in MONTHS:
                    if month in val:
                        year_match = re.search(r'\b(20\d{2})\b', val)
                        year = year_match.group(1) if year_match else str(__import__('datetime').date.today().year)
                        return {'month': month, 'year': year, 'monthIndex': MONTHS.index(month)}
        return None
    except Exception:
        return None


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/detect', methods=['POST'])
def detect():
    """Detect month/year from an uploaded file's contents."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided.'}), 400
    f = request.files['file']
    if not allowed_file(f.filename):
        return jsonify({'error': 'Only .xlsx files supported.'}), 400
    result = detect_month_year(f.read())
    if not result:
        return jsonify({'error': 'Could not detect month from file.'}), 422
    return jsonify(result)


@app.route('/generate', methods=['POST'])
def generate():
    if 'file1' not in request.files or 'file2' not in request.files:
        return jsonify({'error': 'Two files are required.'}), 400

    file1 = request.files['file1']
    file2 = request.files['file2']

    if not file1.filename or not file2.filename:
        return jsonify({'error': 'Both files must be selected.'}), 400
    if not allowed_file(file1.filename) or not allowed_file(file2.filename):
        return jsonify({'error': 'Only .xlsx and .xls files are supported.'}), 400

    b1 = file1.read()
    b2 = file2.read()

    m1 = detect_month_year(b1)
    m2 = detect_month_year(b2)

    if not m1 or not m2:
        return jsonify({'error': 'Could not detect month from one or both files.'}), 422

    # Sort so earlier month is first
    if m1['monthIndex'] > m2['monthIndex']:
        m1, m2, b1, b2 = m2, m1, b2, b1

    try:
        xlsx_bytes = build_combined_pl(
            mar_file=io.BytesIO(b1),
            apr_file=io.BytesIO(b2),
            mar_label=m1['month'],
            apr_label=m2['month'],
            year=m1['year'],
        )
    except Exception as e:
        return jsonify({'error': f'Failed to generate report: {str(e)}'}), 500

    filename = f"Combined_PL_{m1['month']}_{m2['month']}_{m1['year']}.xlsx"
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
